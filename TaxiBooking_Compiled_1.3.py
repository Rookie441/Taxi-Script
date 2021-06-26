import xlrd
import xlsxwriter
import time
import re
from pyxlsb import convert_date
import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.styles import Font, Fill, PatternFill
import tkinter as tk
from tkinter import filedialog
import os
import base64

'''
Changelog v1.3
- Added a file dialog using tkinter
- Removed unnecessary user prompt
- Added combo option (4 followed by 1)
- Added new driver Column J as "Ong"
- Auto-detect file type and disable options accordingly
- Add stream mobility picture

To-Do:
- Add view template beside option

'''

#Program Options
def option1():
    try:
        window.destroy()
    except:
        pass
    global path
    file = xlrd.open_workbook(path)
    sheet = file.sheet_by_index(0)
    wb = openpyxl.load_workbook(path)
    ws = wb.worksheets[0]
    

    day_list = ['Monday','Tuesday','Wednesday','Thursday','Friday','Saturday','Sunday']
    header_list = []

    for i in range(sheet.nrows):
        for y in range(7):
            if day_list[y] in sheet.cell_value(i,0).strip(','):
                header_list.append(i)
        try:
            if sheet.cell_value(i,0)[0].isdigit():
                last_entry = i
        except:
            pass
        

    working_list = []

    for i in range(len(header_list)):
        working_list.append([])
        if i<len(header_list)-1:
            for y in range((header_list[i]+1),(header_list[i+1]-1+1)):
                working_list[i].append(y)
        elif i == len(header_list)-1:
            for y in range((header_list[i]+1),(last_entry+1)):
                working_list[i].append(y)  

    workbookOut = xlsxwriter.Workbook("sorted"+path)
    cell_format = workbookOut.add_format({'bold':True,'font_color':'white'})

    startdate_list = []
    for i in range(len(header_list)):
        worksheetOut = workbookOut.add_worksheet(sheet.cell_value(header_list[i],0)) ####c = ws[str(sheet.cell_value(1,0))] ########RENAME FILE HERE########
        worksheetOut.write("A2",sheet.cell_value(header_list[i],0)) #specific day
        startdate_list.append(sheet.cell_value(header_list[i],0))
        worksheetOut.write("A1","TIME",cell_format)
        worksheetOut.write("B1","DISPATCH",cell_format)
        worksheetOut.write("C1","TRIP FARE",cell_format)
        worksheetOut.write("D1","RETURN DRIVER",cell_format)
        worksheetOut.write("E1","FINAL PAYMENT DUE",cell_format)
        worksheetOut.write("F1","MODE OF PAYMENT",cell_format)
        worksheetOut.write("G1","Albert",cell_format)
        worksheetOut.write("H1","Sirong",cell_format)
        worksheetOut.write("I1","Frankie",cell_format)
        worksheetOut.write("J1","Ong",cell_format)
        worksheetOut.write("K1","Jay",cell_format)
        worksheetOut.write("L1","James",cell_format)
        worksheetOut.write("M1","CHECK",cell_format)
        counter = 3
        for y in range(len(working_list[i])):
            worksheetOut.write("A"+str(counter),sheet.cell_value(working_list[i][y],0))
            worksheetOut.write("B"+str(counter),sheet.cell_value(working_list[i][y],1))
            counter+=1

    print('Success splitting into',len(header_list),'sheets')
    workbookOut.close()


    #Colour Multiple Sheets
    print('')
    def extract_phone_numbers(string):
        r = re.compile(r'(\d{3}[-\.\s]??\d{3}[-\.\s]??\d{4}|\(\d{3}\)\s*\d{3}[-\.\s]??\d{4}|\d{3}[-\.\s]??\d{4})')
        phone_numbers = r.findall(string)
        return [re.sub(r'\D', '', number) for number in phone_numbers]

    print('Colouring multiple sheets')

    path = "sorted"+path
    file = xlrd.open_workbook(path)

    for j in range(len(header_list)):
        sheet = file.sheet_by_index(j)
        wb = openpyxl.load_workbook(path)
        ws = wb.worksheets[j]

        #Lists
        information_list = []
        time_list = []
        split_list = []
        driver_list = []
        price_list = []
        return_list = []
        phone_list = []
        trip_list = []

        #Counters
        total_entries = 0
        first_entry = 0

        #Main
        for i in range(sheet.nrows):
            #time
            if sheet.cell_value(i,0)[0].isdigit():
                time_list.append(sheet.cell_value(i,0))
                if first_entry == 0:
                    first_entry = i #first entry same row as first "time" entry

        for i in range(first_entry,sheet.nrows):            
            #driver
            information_list.append(sheet.cell_value(i,1)) #column here, (i,0) for timing
            s = information_list[i-first_entry][:4] #checking only first 4 indexes
            for element in (s): 
                if element.isdigit():
                    driver_list.append(element)
                    break
            else:
                driver_list.append("")
                
            split_list.append(information_list[i-first_entry].split(' '))


        #Progress
        print(str(j+1),"...")
        total_entries = len(information_list)

        for i in range(len(information_list)): #check elements in first 3 spaces for return/rtn
            try:
                if 'return' in split_list[i][0].lower() or "rtn" in split_list[i][0].lower()\
                   or 'return' in split_list[i][1].lower() or "rtn" in split_list[i][1].lower()\
                   or 'return' in split_list[i][2].lower() or "rtn" in split_list[i][2].lower():
                    return_list.append('yes')
                else:
                    return_list.append('no')
            except:
                return_list.append('no')

            #price
            dollar_sign = information_list[i].find('$')
            if "/" in information_list[i] and 'way' in information_list[i] and return_list[i] == 'yes':
                initial_price_index = dollar_sign+1
                price_index = initial_price_index
                while True:
                    price_str = information_list[i][price_index]
                    if price_str.isdigit():
                        price_index+=1
                    else:
                        break
                if price_index > initial_price_index:
                    price_list.append(int(information_list[i][initial_price_index:price_index])) #string slicing
                else:
                    price_list.append("")
            elif return_list[i] == 'yes':
                price_list.append("RETURN")
            else:
                price_list.append("")

            #phone number
            phone_list.append(extract_phone_numbers(information_list[i]))

            
        #trip list by matching phone number
        for i in range(len(phone_list)):  
            if len(phone_list[i]) == 0:
                trip_list.append(0) #no phone number
            else:
                trip_list.append(1) #1 represents yellow, oneway trip
            
                
        for i in range(len(phone_list)):  
            for j in range(len(phone_list)):
                if j != i and phone_list[i] == phone_list[j]:
                    if trip_list[i] == 1:
                        trip_list[i] = 2 #2 represents white, start trip of twoway trip
                    if trip_list[j] == 1:  
                        trip_list[j] = 3 #3 represents green, end trip of twoway trip
                    #elif trip_list[j] == 2: #potential 3rd trip

        #Edit excel file                    
        column_list = ['A','B','C','D','E','F','G','H','I','J','K','L','M']
        

        for j in range(len(column_list)):
            c = ws[column_list[j]+str(1)] 
            c.fill = PatternFill(patternType='solid',fgColor="5B9BD5") #blue

        for i in range(first_entry,len(information_list)+first_entry):
            wcell1 = ws.cell(i+1,4)
            wcell1.value = driver_list[i-first_entry]
            wcell2 = ws.cell(i+1,5)
            wcell2.value = price_list[i-first_entry]
            if trip_list[i-first_entry] == 0: #gray, no phone number
                for j in range(len(column_list)):
                    c = ws[column_list[j]+str(i+1)]
                    c.fill = PatternFill(patternType='solid',fgColor="00C0C0C0") #gray
            elif trip_list[i-first_entry] == 1: #yellow, one way
                for j in range(len(column_list)):
                    c = ws[column_list[j]+str(i+1)]
                    c.fill = PatternFill(patternType='solid',fgColor="00FFFF00") #yellow        
            elif trip_list[i-first_entry] == 3: #green, 2nd of 2 way
                for j in range(len(column_list)):
                    c = ws[column_list[j]+str(i+1)]
                    c.fill = PatternFill(patternType='solid',fgColor="0000FF00") #green         

        
        wcell1 = ws.cell(total_entries+first_entry+2,4)
        wcell1.value = "Total trips: " + str(total_entries)

        wb.save(path)

        #column E final payment

def option2():
    try:
        window.destroy()
    except:
        pass
    global path
    wb = openpyxl.load_workbook(path)
    numberOfSheets = len(wb.sheetnames)

    total_price = 0
    total_price_list = []
    date_list = []
    total_trips = 0
    total_trips_list = []
    
    for i in range(numberOfSheets):
        ws = wb.worksheets[i]
        date_list.append(ws.cell(2,1).value)
        for k in range(ws.max_row):
            try:
                if ws.cell(k+1,1).value[0].isdigit():
                    total_trips+=1
            except:
                pass
        total_trips_list.append(total_trips)
        
        for j in range(total_trips+1): 
            try:
                wcellprice = ws.cell(j+2,5)
                total_price+=int((wcellprice.value))
            except:
                pass
        total_price_list.append(total_price)
        wcellprice_day = ws.cell(total_trips+4,5)
        wcellprice_day.value = "Total price: " + str(total_price)
        total_price = 0
        total_trips = 0 #order matters
            
        
    #Editing Excel 
    wb.create_sheet("Summary")
    ws = wb["Summary"]
    numberOfDates = len(date_list)
    ft = Font(bold=True,color='FFFFFF')
    column_list = ['A','B','C','D']
    column_list_header = ['Date','Total Trips','Total Price','Total days selected']
    for j in range(3):
        c = ws[column_list[j]+str(1)]
        c.fill = PatternFill(patternType='solid',fgColor="5B9BD5") #blue
        c.font = ft
        wcellheader = ws.cell(1,j+1)
        wcellheader.value = column_list_header[j]

    for i in range(numberOfDates):
        #date
        ws.cell(i+2,1).value = date_list[i]
        #total trips
        ws.cell(i+2,2).value = total_trips_list[i]
        #total price
        ws.cell(i+2,3).value = total_price_list[i]

    for j in range(4):
        c = ws[column_list[j]+str(numberOfDates+4)]
        c.fill = PatternFill(patternType='solid',fgColor="5B9BD5") #blue
        c.font = ft
        #header
        ws.cell(numberOfDates+4,j+1).value = column_list_header[j]

    #summary
    ws.cell(numberOfDates+5,1).value = "Total Summary"
    #total trips summary
    ws.cell(numberOfDates+5,2).value = sum(total_trips_list)
    #total price summary
    ws.cell(numberOfDates+5,3).value = sum(total_price_list)
    #total days summary
    ws.cell(numberOfDates+5,4).value = numberOfDates

    wb.save(path)
    
def option3():
    try:
        window.destroy()
    except:
        pass
    global path
    file = xlrd.open_workbook(path)
    sheet = file.sheet_by_index(0)
    wb = openpyxl.load_workbook(path)
    ws = wb.worksheets[0]

    #Lists
    monthToLetter_list = ['A','B','C','D','E','F','G','H','I','J','K','L']
    clientID_list = []
    work_list = []
    month_list = []
        
    def add2Zero(n):
        if n<10:
            return '00'+str(n)
        elif 10<=n<=99:
            return '0'+str(n)
        else:
            return str(n)

    def uniqueLetter(name):
        if '#' in name:
            return 'I'
        else:
            return 'R'
        
    #Main
    for i in range(1,sheet.nrows):  
        ddmmyyyy = format(convert_date(sheet.cell_value(i,9)),'%Y/%m/%d').split('/')
        work_list.append(ddmmyyyy)
        month_list.append(ddmmyyyy[1])
        if i == 1:
            current_month = month_list[0]
            order_counter = 1
        else:
            if month_list[i-1] == current_month:
                order_counter += 1
            else:
                order_counter = 1
                current_month = month_list[i-1]
        order_creation = add2Zero(order_counter)
        clientID = ddmmyyyy[0][2:]+monthToLetter_list[int(ddmmyyyy[1])-1]+order_creation+uniqueLetter(sheet.cell_value(i,1))+add2Zero(i)
        clientID_list.append(clientID)

    #Debug messages
    newID_count = 0
    correctID_count = 0
    wrongID_count = 0
    wrongID_list = []

    #Edit excel file
    for i in range(len(clientID_list)):
        wcell1 = ws.cell(i+2,1)
        if wcell1.value == None:
            wcell1.value = clientID_list[i]
            newID_count+=1
        elif wcell1.value == clientID_list[i]:
            correctID_count+=1
            pass
        else:
            wcell1.value = clientID_list[i]
            wrongID_count+=1
            wrongID_list.append(i+2)

    print()
    print('#####')
    print(wrongID_count,'of',correctID_count+wrongID_count, 'existing clientID had mistakes and was corrected.')
    print('Rows of initially wrong IDs:',wrongID_list)
    print()
    print('#####')
    print(newID_count,'new clientID created.')
    print()


    wb.save(path)
    
def option4():
    try:
        window.destroy()
    except:
        pass
    global path
    file = xlrd.open_workbook(path)
    sheet = file.sheet_by_index(0)
    wb = openpyxl.load_workbook(path)
    ws = wb.worksheets[0]


    day_list = ['Monday','Tuesday','Wednesday','Thursday','Friday','Saturday','Sunday'] #Must be Caps first word
    header_list = []

    for i in range(sheet.nrows):
        for y in range(7):
            if day_list[y] in sheet.cell_value(i,0).strip(','):
                header_list.append(i)
            if len(sheet.cell_value(i,0)) >= 8:
                last_entry = i
                
    working_list = []

    for i in range(len(header_list)):
        working_list.append([])
        if i<len(header_list)-1:
            for y in range((header_list[i]),(header_list[i+1]-1+1)):
                working_list[i].append(y)
        elif i == len(header_list)-1:
            for y in range((header_list[i]),(last_entry+1)):
                working_list[i].append(y)  

    #Clear
    for i in range(sheet.nrows):
        ws.cell(i+1,1).value = ''

    header_counter = 1
    time_counter = 2
    trip_counter = 2

    for i in range(len(header_list)):
        
        entry_row_count = len(working_list[i])-1
        ws.cell(header_counter,1).value = sheet.cell_value(working_list[i][0],0)
        
        for y in range(1,entry_row_count,2):
            ws.cell(time_counter,1).value = sheet.cell_value(working_list[i][y],0)
            time_counter+=1
            
            
        for y in range(2,entry_row_count+1,2):
            ws.cell(trip_counter,2).value = sheet.cell_value(working_list[i][y],0)
            trip_counter+=1

        header_counter = time_counter
        time_counter+=1
        trip_counter+=1


    wb.save(path)


def option5():
    option4()
    option1()

def availOptions(availOptionsList):
    if availOptionsList == ['disabled','disabled','disabled','disabled','disabled']:
        return "red"
    else:
        return "green"

def checktype():
    global path
    availableOptions = ['disabled','disabled','disabled','disabled','disabled']
    wb = openpyxl.load_workbook(path)
    ws = wb.worksheets[len(wb.sheetnames)-1] #Last sheet
    if ws.cell(1,1).value == "Date":
        #Summary file detected #No options available
        wb.close()
        return availableOptions
    wb.close()
    ws = wb.worksheets[0]
    if ws.cell(1,1).value == "Client ID":
        #Passenger ID format detected #Option 3 available
        availableOptions[2] = 'normal'
        wb.close()
        return availableOptions
    elif ws.cell(1,1).value == "TIME":
        #Sorted format detected #Option 2 available
        availableOptions[1] = 'normal'
        wb.close()
        return availableOptions
    if ws.cell(2,1).value != None and ws.cell(2,1).value[0].isdigit():
        if ws.cell(2,2).value != None:
            #Standard format detected #Option 1 available
            availableOptions[0] = 'normal'
            wb.close()
            return availableOptions
        else:
            #Online cal format detected #Option 4 and 5 available
            availableOptions = ['disabled','disabled','disabled','normal','normal']
            wb.close()
            return availableOptions
    else:
        #Unsupported format detected #No options available
        wb.close()
    return availableOptions
    
def decode(imageEncode):
    imageDecode = base64.decodebytes(imageEncode)
    fh = open("Logo.png", "wb")
    fh.write(imageDecode)
    fh.close()    

#Tkinter Interface
def openFile():
    global path
    directorypath = filedialog.askopenfilename(title="Made by Rookie441",filetypes=(("excel files", "*.xlsx"),("all files","*.*"))) #C:/Andre/testing.xlsx
    
    #Using exact path rather than relative path(working directory where script location is)
    os.chdir(os.path.dirname(directorypath)) #navigate to C:/Andre/ 
    path = os.path.basename(directorypath) #testing.xlsx  but since already navigated, it is actually the full path C:/Andre/testing.xlsx
    
    mainlabel = tk.Label(window, text="File Path",fg=availOptions(checktype()))
    mainlabel.grid(row=0,column=2)
    mainlabel.config(text = path)
    checktype()
    #Buttons
    tk.Button(text="Option 1",command=option1,state=checktype()[0]).grid(row=2,column=0)
    tk.Label(window, text="Split into multiple sheets and colour").grid(row=2,column=2)
    tk.Button(text="Option 2",command=option2,state=checktype()[1]).grid(row=3,column=0)
    tk.Label(window, text="Tabulate total price, create summary sheet").grid(row=3,column=2)
    tk.Button(text="Option 3",command=option3,state=checktype()[2]).grid(row=4,column=0)
    tk.Label(window, text="PassengerID Generator").grid(row=4,column=2)
    tk.Button(text="Option 4",command=option4,state=checktype()[3]).grid(row=5,column=0)
    tk.Label(window, text="Online Calender Formatting").grid(row=5,column=2)
    tk.Button(text="Option 5",command=option5,state=checktype()[4]).grid(row=6,column=0)
    tk.Label(window, text="Combo").grid(row=6,column=2)

    #window.destroy()

#Window    
window = tk.Tk()
window.title("Made by Rookie441")
window.geometry("350x350")

#Display
tk.Button(text="Upload File",command=openFile).grid(row=0,column=0)
decode(b'iVBORw0KGgoAAAANSUhEUgAAAUAAAABxCAYAAACkwXoWAAAABHNCSVQICAgIfAhkiAAAAF96VFh0\nUmF3IHByb2ZpbGUgdHlwZSBBUFAxAAAImeNKT81LLcpMVigoyk/LzEnlUgADYxMuE0sTS6NEAwMD\nCwMIMDQwMDYEkkZAtjlUKNEABZgamFmaGZsZmgMxiM8FAEi2FMk61EMyAAAgAElEQVR4nO29eXCc\nd3rn933v7rcvHA2gCRAESY14iJREieIclEFNLHtJ2TNaS0p2t0acWu/GI83WlsuUK5ukyEm2NhFV\n2axrqNi7qSEnmY3HlGvtDUVbGkdkImkkcihrRFKkJJ6geAAgcXU3Gn29/d5P/njxNrsbbwNooHnq\n/fxDoo/36H77eZ/fc3wfBvc5mmnRaNGCpVswiBAJ8OiJisydPi4fH587z31rCI4MFSiT0TCl25BF\nFhzDAgAssiGwDKIBDk+uCEPiufv2M/Dx8Zkd9k4fwK1AMy0aSalIqzYYlsNAwcK5rIHzeQMcw8Im\nBhN5A78eKkAzLbrTx+vj43NnuO+8n89HinRurASR53ByUsPVvOn5ujDP4HeWBtEdFvD1FZH77nPw\n8fGZm/vOA7yS1lDULZzPG3WNHwAUTMLlnImCaWOyaPheoI/PVxD+Th9As8mqFlplEe8NFud87clJ\nHcMl6zYclY+Pz93IfecBWnZjr58oWSg0+iYfH5/7gvvKAN7I6SRxwGdZvaH3PRwP3qIj8vHxuZu5\nrwxgT1RkTALiQmPvs8kPAfr4fBW5rwwgAPAMkAjOP7QpCyyKhr8E9vH5KnLfGUBRYKHo8zNoAstg\n6xIJ0VCDLqOPj899wX1X/5YsGnT0cgEA6tYBdgY5rIkJWBHmYDHA76xpve8+Bx8fn7m57zzAjpDA\nPNoTRF4zsaFVgMBW27YVER5bOgPolDgEeBaPdofu0JH6+Pjcae5bz+dcSqXhZAklk2DTzdMMCcCK\n1gBG8zqeejB2356/j4/P3NzXBiCr6fThlwW0Szw4lkFIYMHzDEYLBlZ3BX1VGB8fn68GWU2nGznd\nr3fx8fHx8fHx8bntS8CibtHJsRJ+fi4Hpab+bkuvjB9uaLtlx6SZFukGwSKAY4BIkPeXwD4+X2Fu\nqxjCRNGkP3p/vGz4OmQORYPKf6eL9dVbFopp2VTQbGimjXTBREaxIHIMelrEpu/Lx8fn3uK2eEC7\nj03Q2UmnP/eFVREcGMjjhVURvLA6xhARvX48jePjavmxZuyzoFpU1C1YthP2u5RUcTGlgmOAf7qp\nAzzH+t6fj89XnFvuARIR7f0sgxWGjd3fTjBvENGnoyW8c7WIPzg0QoeuFNDfK+P4uNqU/WmGTemi\ngT/91SgebA3gckaFPm0ENRtIaSb+id/55uPjg9tUCN0R5CALzq6SioWC5Sx7FcPGp2Ol8nOLxbRs\nSuYNHLucx4qYhGTJKBu/Cc3CtbyOB3tDCNVWR/v4+HwluS0xwI2JIA4M5LHnkxTteH8MALCuTcTW\nlWFsTNyUorqQ0ha8j3TBoLGsgXcvTuFURkU77xjVKd3GeMnEmAB8xph44/FO3/j5+PgAuE0eYF/U\nERt4bnUUW3plAMDZSR0/PjGJo8M3lZtnk7Cfi5xq4fPrRZzKqEjqNiwCBkvmTeNnNj/B4uPjc29z\nW3uBFcOGPF158sZ3erClV8ZPPpsqZ4Fry2LmS1G36PRwAV9mNYRZFgmexUcTChJhAQMBYKnIAQD+\n1dr25pyIj4/PfYGnAZwomk3rmGAYprzkTCkmTo6r5Zjfy4+2AsCiYoCKZtGlsRJUkxDlWRRsG5c1\nE092hyAR8LuyhI8NE+0Chx9+LbrIs/Hx8bmf8IwB7nh/DN97+zr95XeXNjVedvhKAS8/2gpZYLHj\n/TEkFWcgkfvvQvj7y3lopo1LWSd+eEO1sD4kYiBv4JhtoFPgkDYspP7L1UysKWfh4+NzvzDD9SK6\nNfrwKyI8ruZN7P0sg51HJqqM3mt/n1zQNjXDJs20ca2goWDaOJnV0BPgMFJ0jN8G4nBV82N/Pj4+\n3swrC1zULXrto2Q5SdGoZ8gwDDNRNMn1+jZ1BdDfK+PEuIqH2kT094bKHuFE0aSj14vYmAhieay+\nWsuUYtKvLuXwwUgBedNGUrcR4lhYFmGIIzwJAX+wrg3HPhvDv3m4A/+ykQP28fH5SjCv4NvOoxNV\nGdpr2cZVVTpkDlt6ZWxbEcKOTe3YmAji5UdbceBSHq99lERnwElUhEQWn46WsPPIBP7ovVGqt6+J\nnFH+fxvvvDfBs7AYJ+P7G50BXNdstAsc/qulfuzPx8dnJlUG8MDFLFUmLY4MFeh7b18nd7n6o2/F\n8cZ3emb1zOrxg8OjeP7BKB5qlwAAR4eLYBiG6QxwODup4+ykjj9+og0hgWVefaoLssAiqVjYeWQC\n51JqlRG0bKLxnNNaJ4DBmaKOByUOGc3Cf1adWOD6hIz/YSCJP9+0BCXVXwb7+PjMpGwAT4wqtHba\nOLn85LOp8v9fWBVB7fON0BVg0SFzGJz23H7y2RS+9/b1co/wighf3h/DMMyPvhWvuy1Fs8CyDCZL\nJrKGBYFhMGnamIRjJ1dIPH4xoaBoE3rDAjTDlwH08fGZCQs4nt+PT0zioXjA07N7YVUE/UtDeO2j\nJE6OlRraARHRH703Ss+tdpahBwbyUAwbP926BCsiTgjyR9+K49WnuvDMihB+cnqSAKd4+o3v9EAW\n2KrjKukWDWV0FFQLN/I6xgwb60ICxkwbx8jx9DZzPH52PQcAmCqa6Iz6U998fHxmwgLAO1eLdV+w\nIsIjWXKWwGcn9YZLVo4OF5FUrCrv8eR0/+/OzR344aMtZQO3bUUYR4YVnBhVykvxn25dUrU9zSSo\nhoWCbqFg2OgQWXAM8CWcIuoQy6A74BjWdoEDCAhL993sJx8fnybAAtUdGJVJhze+04OdmztwZFhB\nhzydpBAYXMvqtPvYBP38TIaKujXr+tJd1oZEjmEYhtnUFSh3f4REjtmyLMwA1QXTPz4xCfexyscB\nIFsyoRqEqzkdScNCp8A5/b7knMPXWA4dEcfj++24jHiY96WvfHx8POGJiM6nNRy86CwZ3zjjGCy3\nO8P91+3ZTZYs/PxcCk8kAjh0tYgj1+sviYmIXvzFjarHnlsdxfFxFT84PIrdxyZoTdzxDC+kNPzg\n8Gj5de9cztMzD0RmGK6CdtNY520CZ9vIWjcf6+IkTJScpbDMMWAZ3/b5+Ph4wzMMwxARrd3cAVQI\nl4Z5Bns/y5RVml1P7kJKq/IYFcPGRNGkzhDP/OT0JD3RFcATS+Sy1dm2IoRDFUtsN7b34i9ulLO/\nXvxquIiiblFI5Kos2JRiIqs4x5TWLCyPiDicv7mNAGPiStHxANcEBYREf/nr4+PjTVUhtGuMOmQO\nE4qFPt2uMlAdMlf++8SYije+0wPA8Q53fTBGR4YVnBhTqwzX99e14IVVUfzl9DZcg7ulV8aRYWXG\nAW3pldEXEbBtZRi1y1/NtOiTKwWUTBsZ1YlFlgwbBYaAioV4VCKgCExavvKpj49PfXjAMUoHLmYJ\ncDK+RcPGoatFbF0ZxvFxFa9t6USHzOPQlTwODOQhCyz++Im2sjEDgMeXBDGuFqEYdnkpW2vAXNz3\nvfxo64zX/KXXG6YxTCCvWigaNiZNxwDesGwU7ZvWb9C08M2gozH4Z9em8L2Vfgewj4+PNzzgtLr9\n0fvj6JA5PL8qikNXCgCAeNBxEH96ahKywGJi2uvasjSIgxdz+N7b1+m1j5J4cX0L+nuF8ns/nUep\nTD3jOBtum7I9bfBCHItRHkDNKlrRnOMs2oSi5nuBPj4+3vAA4E5q++nWHjAMw7iZYFe9+QePteGN\nM1NIKhY6ZA7fX9eCk2Ml7NzseFpuH++6NhE7N3dg28owdjV4IJemJmlEKWJDexdikneniW4RjGkj\nmDUJbRyDcb26LOc6Q8iZNr7ZEsDHUyrSRRMjUxp1t0h+NsTHx6cKFnAKkV2lZuCmgjPgZIFPjpVw\ndlJ3+nmXymAYhvmLc9lyMuS1/k6siPA4O6lj72eZhoRNPxwZpp8PnKU3h6/i4/QEjo4NY1It0ZXc\nFB0bvU6pkkKTasmz1MZigBxX/VjasCCKLL4Tjzh/lwxkSwuX2/Lx8bl/YYCbEljusvTIUIF+8tkU\nXtvSCcWw8erfpwCgnPT4weFRKIaNH30rjofigXIc0C15WRHhsfvbCU+PK6vp9PHEDXw8OVF+TAUw\nZppImgambMdY9cthBKaf75SC+BdrHmHSBYN+dTmHjGLi/aSChMjhYtEod4BU8j91R3HSsPB5TsOP\nlrXgqQdjaJH9Qeg+Pj43YYGZ8bj+3hBe/83EDE9w72cZMAzDrGsTnb7erNPXyzAM8+ZArvy6HzzW\n5rmzjKrS6xdOVRm/42oJR5UCLukqpmwLCV7Aw2KgbPxcNLO64FogIG/aKNR4gC43ciYkg3BVM3Es\nryGjmNAM228K9vHxKVMug6k1gh0yB4ZhmD2fpAhwOkDSRRPfe/s6vbiupfy8q/N3YCAPwGldq93W\ntdwU/fnVi/jTi5/hw5JTE6gTQWQYyAyLB8UAVgWC6BEk/HUmiSSAVUIQYU7A5o4lCPHOYUo8iwDH\nQuJZEMfAtghddRS9/qao4vdCjhm9ppv49VABDydkz9f6+Ph8NfG0HpUtaC891oZ1bSKSioUdm9qx\nIsJjx/tjeO2jJHYfmyiPuQSAHz7aMsP4HR6+Sn9+9SKyloUPS0XoRNCJ8KAYwFPBEDYFgmhlGPxd\nLoO0qSHAsoiwHJYHw9jY1oG45BgxiecYgWPAswwElkGcZ5EnIMg4/b+1jJONYd3Eb7YFkVEtnCto\nuJRWMaU0b96Jj4/Pvc2cMTEioqRiYcf7Y/jp1iXlpMjBi7mySOoLqyJ4flW0yviNqiXaefZT5G0L\nEZZD0jTQwQt4Qg6DZRi08AIigohWUYJJBNUywYGBalsQWBYcGDzcFkdIEBEWb2aFj32Zo6xiYqJo\n4OCIU67TLnF4S/WeKbwjFsLqziAGLQvn8zq+2ybj6a9FEQne3njgiVGFDl8poC8mYPv6Vj8W6VOX\nl965QZWJxJc2tMLtmb9X2PNJihTDRn+vfFcf+5yS+JXFzn9xLouXH23FxkQQGxNBnE9rGMwa2LYy\nPON9f3rxcwwaGjaIAQxZJh4OBNHC8ohwzi6XhSLgGAYWEXiGQbsUQEyUILAsJI5HTHR6hGVBqPrw\nohKHwnQ9YrfEIanbiMyidfDLogYhzYBrERDkWZzNmUiMlpAuGNQeFm7LF3Mtq9OPPhwHAJxPazhw\nIUsvrIndtReFz52ltorizenwUiMkiwZ1hGa/vt2bsiyw2JgIVBmqZNEgWWBR24o6X1zZvPNpDbuP\nTdCuJzvvyut9XjNBAGBTVwBHhhUMT+nlDC8R0dp2qcrzu5BJ0V8NXcaHpSI2iAF8opXQL4exNugY\nya6gDA4MwoKIzqCMmCjNMHKzkWgRkdMs8AqDh1skXCsYuKLUV3z+zDTx2WQB/z0bwaaQgP9xOAMA\n+DJVwsBYiXrbRAQX+CXPF1dgwuXgQO6uvih87i46gnUyfXVIFg165b0xbH9rmAAnMfnShtaqHn0A\neP14uvz/k2OlKgGSV95zQlvuNhrxQk+MKlS57fNpDedSKtXTG72TzEspgGEYZlnMyQhXzgbxkqv6\n2+uD+EJ3PL8MUZXxiwgCAhyPFklCbziCjoDckPEDnJkhLTKPkMBC5lhkNQsGEbqY2U+lrVXEQFrF\nv36gDfvGpzCgGvh0pIhPh+prITaLwYr5JT4+jRKX5+2nAAA6QgJTOWtbMewqY1ePoYrrtHZW977T\nGew9NTmv+LlbHVLJYmZ/30rmfVTPr4rije/0lGsBvfibwcv0dmEKY6aBT7QSEjyPtcEwIoKA7mAI\nS4IhrGttx4b2LrQFgsxCdPpkiWN6WkRIHIuwyCE9XR3zW0HRMxni8t9eSuOSSBgcK+HfLm/DJMfg\nXw9P4dNUCW+cSNKHl7JkWremTMarMHyrR9jAx8eL51dFGn7Pvmd6ZvwYXG+uHvEKT3PfMz3MSxta\nq553JfHm4uj1mSInC5kjdDtoqlk+kh4v/9/1/NqkAGReQIskISGHEA/KCzJ8lQQFFhzLQBZYrJy+\nOyZ1G08Ks5e5vD9ZwgM9YfzqRgGPizy+1x3BMGyQTRjOaPh0qIh86fZkiRczX8Xnq8Vcsbx67Nzc\nUeV5NepJblkWZhp9T1G3KDVLSOpuY95nN5d4wU8GL9H+kSEAwINiAF+PtiLCiwCAJaEwVsfamqbM\nLAksc3ZEobEpHZs6ZAAKLhYNCLaOdoFD2qjf+vbHFybwz5dGkYgI+PRSFpsSISxrd47zSkbDQLKE\nw+cztKEnhK7orblrPbcquuDgss/9TWXCbLE0I+b2+m8taWgbXhJ3jRrR20lTPMBRtUR/PToMAEjw\nAr4eikDmBQR5Hm2BAOJSoOmy9PEwD5FnEBJYrG91agUNIjwaEed878+u57DnyhRWdwTx0UgB+ZIF\n0wI4jkE4yCOZN/DLL3O3rHNk45Lgrdisz33AveQ9eTHkEe++m1c7TTGA/+7cKXwjGEKCF/D7nT2I\nixICHA+J47CutQNdoebXAXVFRebhpSFIPIt4kMfGmASBYaBlDPwDfm4jeErR8W+uT+FT1saHwwX8\n1blJXEmrSBcNuGbv//4sjTc/z5Bl04IN4YlRZcZ779Z4iM+d56iHB3W7WBtfvKHyihP29969HViL\nNoCWTdTGC7ika3hwuo6vVZAgsSxiogSJv3VLvUiAQ0B04oHLQgIeid40fN9sqe0m9uaqZgJhHt/o\nDmEgr+PLtApiGRR1C0GRR0k3cfy6I8+/kGP0yoj5+NSjESWle4W7sfzFZdGL88/S4/i4VEQHL+CR\nUNTx/FgW7YEgVkRamnGMdeFYhpksGjQ6bWRCIgfYhAuKiYdMBl+PynijpM0aEwSAn405haZ/2NeC\n/pYA3jg3CdskbFpig2dYXJ4o4fJECdcnNVraNlNXsKhbdGRYKV+8lXfS2hpAwPEKa+cdN0qyaNCb\nA/kZd9xXn+ryPcx7mPPp6o6m27l8bKTe8MCFbLnGsC92UzTltY+SVa+LyzzeuZynJxKBOZM551Iq\nHbyYu63dUovayWixQH8zfAWfF/NI8Dy+2daJmCCCZ1lsiHeBZ2/POErLJvriehF51SpL5r8/WsCI\nZmF9SERHgMO7poEvclqVfH49/mFHCP+oO4z/9PkkWJ7BAyEBMs+iNSSgVebw7Lq2qvPafWyCai/c\nhbAxEcQrX4/P+pldy+r001OTs9YW9veG8PJj1ce4/0yGToypM2JM/b0hPL8qMufFeWSoQCfH1HKF\nP+Bc3H1RAdvXxRrKVJ5LqQQ4noFX0H/vtu6qJNG1rF5uDVusN1G576Ju0cuHRua9776oUDd5dS2r\n0+vH0+XPNy7znvE8r++mktpSlb6ogKJJntuKyzy2r4vNKHCeD16f+/5ne+e1nQMXsnRwWv1pMcRl\nHru3dCIkcozXb2jHpvZZz63ys5IF1rP0Zy4W5QFOaiomtBI6WAEcw4CbtqcBjr9txg9wPMGr0xe2\nRQSOZfB4awBiVsOYZmJYN/HtqITv9IawZ2RqTo/wb5NF5C0b/T0hfHqjiMGigUSABzt97U8WDWqr\n+MH3xYQZd+6FcHKshL2nJmm2H8hsxq8vKmDrynBV1q2oW7TvdKY85qCWo8NFHK0zgQ9wDMbhKwXs\nO52Z8d6UYiKlmDg5Vmqos8X1EvaemiSvAt3XPkqWj+fAhWzVD/Wld27QQi50r33Xeiu1+z4yVKja\n92zFvJXGD6ifzDg6XMSJUYXma7Rmu9GlFLNugXOyaFCyZKEjyHnenOZaas/2/i29MpphAFOKiVfe\nG0OyaND+s9kZzx+uc80CznVdefNSDBtHhgrUaN/xoozULwYv0/81cQMPixIeibaiTQqiRZTQGZSx\n5BYkPubDF8NFKho2iqoFzbRRNGzcyOtIqhbGNBNBlsFD7RJyAos/GZy5PK0kxDL4x90RpMdV2Cah\nOyQgKrJY2R7AAy0SHu8L3xJP0GsZey2r02sfJWdcuPXanNz3NFpS4XoB9fY3Fy+ua4HXPGeXdy7n\n6Y2zs3/ucyELLPY8nWi4lOjIUIG8DHmj1HqJCz2n2u2cS6meRnk+VHqWtWIKlZ6W+1htuxpQ7QHW\neqK1x1rULTp0pdAUQ9jfG8L2dTG88t7YjOvNyyt1W/28mK8X67IoDzBv6FBtGwGWh8wL4BkGEsci\nIsydhb1VPNAZwHjewIhN4KZDGqvaAhCmNEgsMKFZOJlUsTTA44+6ozgypeKU4j2buGgTfnY9h/9j\ndQfevpJFdtpzLKkWrhd0WDYRVzF5fdeTnYwbmwsJDLyWnBsTwXIXyPmUYywvpDUkSxZSiomXNrRi\neUxk3CUVAPQvlVHPGO3c3OEZ83MN2EJZiPEDgMNX69+1Ae8yiUZRDLtqKT5fmrFvoLrWrahbdX+M\nwM2b2YlRhQ5ezFV5dI1+P7LAYvu6GGSBxf6z2aprqzIWXPu9uV56JbMl57wMTG19n2sMjwwVaChn\nIMizSJWsGTHpHZvaIQssFMPGYNbAUM6AYtjlFZNr/EIix+w9NUnz6TapJw7R3xvC/jnfXc2iDOA1\nrYQAy2KJFAAHBhzDIMRXy1fdbmTJ+WJKukUTeQPjOQOGRVjDszBtgmUTBiZVFAwb5yZUdAD4V21h\n6AEW/9uI993sDy4m8eM1nTh8fhIZ1UJUZGEQYWhyprdXuVx46Z0bMwKOL21ondVzcb/ASs/N6y7b\nFxWwc3NH3W3V8/wq7+ReMbB6j8sCO8PYenksc9WxNcNDBhZWLtKsjHzl1MNdRyY8bxSueID7fboe\n+q4Pxsg1goM5A9vfGqb5eC1ujGtfxWO1Xtq1rE71EmBJZf5zcZIeM3Tq3Qwrl5xuYqSSuZb6+wG8\nPP3/lx9rY2rPaf+ZDNUmROoZydlCR/VYcBmMZlo0oJbQwXIQWBbW9GRykb07mp6DIscsbZXQFRXQ\nIvMIiCzk6ZKZVW0BrGkLYH1YwNIAh+NTGq4lVfx4TScek7291z+7OomeiJPtUgxnaT0ypUPR6pfH\n1F40cZlvWgfIbMYvWTQ8j6m/N1R+z7WsTruOTHi9zNO7qudpNko9A3k7muXrGd/F7Lv2fOIyjz1P\nJ+oqp7y4fmGVEc+tis75mtk89toavwsN3ojm8xk1us35cOhKoe713AwW/M2rlnOXSPA82qSAm/hA\nPBhq2sFVYlo2pQsGnR9V6PyoQumCQbMZH8BJjvS1B5h13TLz+LIQViWC6IgIiMk8lsREbEyE8O3u\nMP6LeBB9QR7vXpnCYyyHV/va0S5UlwRc1Ux0dgbweHcEQwUD40UTl9MqzozM3xPpX9qcgtAX17XM\nakh3Tw+xqmT/s73My4+1MUeGCrTj3VH60YfjM368bsmFl3d1q0trdm/prPtcbVN+s9m5uaPuc/Mx\nPJW8tKF11oy4VxZ7rhrTnZs7Zo2rzrbtZlFZ6uJFsmjMiH8v5Mbi9V1ULsd3H5vw/KwWeo0s2ABa\nZCE/PcGNAwN+OhRWGRNrJr++WsBPP03iJ6eS+PC6gsmiic9vzN/4BEWO6YqKzPK4hKjEQeAYhAMc\nZInFihYJ6+JBfC3AI2/amMiU8N91z7zw/+zaFB5sF6DbhJJpwyBCRp1/61KH3Jiumxd9UWHWH0Oy\naHg2o7/0zg3a/tYw7Tud8fTC+qICdmxqb+hYGl1SuiUoXvuezWgsiy0+plzPyMRlflbjvuU2dDHM\nJZd2NxcSu3idQ+VQtfnyUDzAbEzUbxX18uLjMr9g1ekFG0DDsqDajsttgcAxDIJ84ydcUC36+Eqe\n/sORUfoPR0bpTz4Yof/9o1E6O6JQQbXoHx64TA/tPUNnRov41WAe/99wHn/66RjePDuJsaKB00OF\nhtzjtpDAPNwbYr6+IsIsbRXREuTRFhbQ0yLhG91h/IOeMCwC3r1RxM/XVy+JizbhD885Ma/xklMC\nMlYwUZpnl0gzpMHrjRt18YrfALMvj7atDGP3txNMo8vzekueRpcssxleWWAxlJ2ZpFrTYIFwPSMz\nl9H3/GHP4g3NdVPYfyYz47NxDdxD8QBTq65eT0ig9mbSqLfVrFisy8kxdcZjj89iyGbDqxZ297EJ\nqncTa1SwoZIFG0Cvn9NColtvfzGJvz0/iXcGc3hnMIePRoo4l9bw11+kcXK4gEspJx61dyCDK8rN\ni+vPL2WQLhoo6va8DVAtfe0BZlVXEIHp8SDhAIeoxKNV4vBgkMexawX84crqmE3RpnIsUDNtmJaN\nnDrT6FzL6vfE8KWdmzsWXHVfLxN7wuPHMBuzeX9bV4QbCuA3ymzeX39vyNOgBfn6P5u5Muf16jFd\ntq9vLUtQuSVOtbi1nZXM5jXdDjxXFXMsmxvhfFqbl6hroyzYAE5qKlYtstzlZ8cn6PiNAr6YVHFF\nMXBFMXAhr+GzSRVfpFX81edpPNQhY01EgmrOvLCYaYu7mFW3LHHMhmVh5rFlsqM2HeLwcKeMh+NB\nXFQM/J9nJ/GbbdUX15olzrKIZxlkVAu5kjVjbrFb4nK78VpCyAKLbSvDWNsuYW27hOdWRbH/2V5m\n/7O9jNfyaj5ire9cztc18G+cnfL0Auf6TGrjPzs2teOFNTHGy9NsdGk6175ffaqryot6cV0LXn6s\nzXPflWo+tcu8gwM5z3M/cCFLXoKkXvN0Xv+tJcz+Z3uZfc/0zPh+zqVUevnQyAyDs31drP7JNchC\nYndeHuVilu57t3XP8H699vHiusW12y7YAIosixhXHdPi5pClr2WiZGJUMVF7WmnVxIW8ho+nVFzJ\nqLiQ9754YxwLe7q0ZbFYZGNJi4iA4AgsRCQOq2UBw1Manu2sVuS9UDIQFlloFkE3bRABRs0N8E42\ntdcaMMWwsX19K7PryU5m15OdzFwDmbxKFyoN3p5PUnMW/nrpwnlR+WN7KB6oUiGuV0IhC+yCRULr\nsTwmMpXL4dnirJVeo1dWt7ar4VxK9Wwd64sKnkmWom6RlxHdfybjWdtZmd1vBs1IeC1WAzAkcsxc\nRl0W2EXHaBd8lO2BIAIsD9V2fvkWEQSWhWnbNN82uLxho8zG2wQAAB2dSURBVFXicHHKO27i5fW5\nPPe1FtzI6vjG8ki59m8xuLNJTMumi+MqBI7B+vYAUpqJz6/n8Yd9Lfiz6c6RXyYV/JN4BDcyJXAM\nC8MmFPXqZdqtKAl4blV0XoWeD8UDTGW9GVA93CYu83PenfuiQlX8642zU+VtzKcI+eBADjveHaUd\nm9rLPygvI7B9XQyVtW3ziZPWvmc+eBVov7iuperznI/H4vWe2rrJk2Olqho9L6PVFxWw+9sJZrfH\nPhTDRuVQIxev5bMssDPq3/p7QzNq5Wq9Orc4+VbR6CAnL55YIs+4jitZSDdQLYvwAG+eoGHbMIlg\n2HZD3lhvNIDlEQHtYuMf1oOygGiAQ1uoOWqz40UnmeIKt4YkDgGeQVzioRg2VnA3P6qrmomOIAO9\nQibQqFHSr5eMWAyN6LXt3NzhmYXbdzqD1z5KYvtbw7T72AQduJAlL83CHzzWNu+l0KtPdXmqlqQU\nc85uh/mondR6E82Kdz2RmFsyrTaO5eVxhESuIen4te3SrKU3tUONZtuOV/nQRo/zqvXqaq+NuY5/\nIRndZuB2knjRDK93wQYwwPNw91+yTBiWBYtsGPb8f/g/3BRnelsD+O0uGb/ZPf8hQT2y82U80h1q\nivcHAF/mbi7pElEBLANEJR4dAQ5F3cZEzV3ocnH2u+diMlOAU7dXGd/YtjLcUEwlJHLM7m8nmNku\noPNpDQcHcnj9eBrb3xqmyszi8pjI7Humh+nvDXm+Py7z2LutG/uf7WWWx0Rm15Odnj/aSi9j/7O9\nTOUP/7lV0XktZV9+rK2cHV3orNp9z/QsaN/b17dWLcvr7duN281muGSBxd5t3dj1ZOecGfc9Tyfq\nPvfShlbsf7aX2fVkJ+N1Dk8skauO2WuZvevJziqjXZtsqc2Oz3W8L21orbpOmpUA6QgJjNfxN1qf\nWY8Fu088yzL/6eoA5YwMTPvmRa43YAABIG/aiAV5pFMqfrM7jPdHZs+SAcD3V7WiJyKit615WmkF\n86aBk3gWgekvU+S9v/f/Z6qAb2B2z/XVp7pw8GIOJ8dK816+VvLMAxHmyFCBkoqFhQ5Sf2KJzBR1\ni3YdmViQ3Lq7vDqXUqkvKuDQlQI2LglieUxkXq957c7NHbOq1QDOkvGdy3lSDLuhc9q+vpXZfWyC\nHk8EG17+Vu57IZ/nlmXh8r7n+g73PdPDuAPHK9m6MownlsjMfI89JHLMuZRKn46VypnoxxNBzEdX\nr/KYAe9EC+AUn7vL6tqb6xNLZGbPJylSDBt9MWHO896yLFx1nTXLQAHArzyW8/XOqVEW5aUM5bP0\nH69cQFwMoE8Oo1sOoUcOI9LgsPOPhwr02WAeF9MqdI7BRxNK3fjfv1zbjrYgj+8+3IZwYGHe33ix\nQJXHeCGTIsU08XiHU2OnGTaliwYGxks4MargVLKE5RERewtOYD/EMvhvumO4mlLRHRbwu6tb0Bbi\nb9kQpXsNV0PvXijg9bm72fNJimpjznueTjQtCbaoBkyJ5RAXA9DJgmFb0C0Lum2j0VlC31wWZjrD\nAh7pkmHrNp7p8Z6D2iMLaAvyWNYqLdj4AUDO0JHXbyYpxkslyPxMZ9iqcGYrHcFVAQG6xSDIs2gN\ncLDJKYnxcVgeEz3La3x8GsUr4dbMCoBFGcCoFMC3u3qQMwykdA0WEYqmDsX0lpeajeceizP/6PE4\nnl4RRZfEoT8RwrcSN/uK/+mDrfi9vig2r4zgieUL7ze2bCKLbASmBycphkETqoI1rTerzw2LYNtA\nUjWhT3uiEenmcveHK1twZqyA1gCHpWERAYGBLN4dIhA+PvcLXq2TzR4RsKhfbZDnmbgUQNayMKnr\nKJlOjEmzFjbaT5Y45hvLI1jXJaNP4mGbhO/0ObEES7exokVCIiouatBSRnPuKIGKGsbuGgEH3XK8\n2JxmwbQJIsugMgZcVEzoNiEh8+BZBjzLQJylO8DHx6dxvBShn1vdvNgi0IShSF2hMPPPTh6jmG0i\nb+qY0jS0SE6cLR4MNSyO4A4dShcMupxU8fG1PH5/RQy/90gcS1tFSMLi1popVcHSUBQSzzGmZdPp\n9DieXLK0apuKZkPRbSiahQnNQqvEQQk4Bi7EMjh8OYewyKJDFtAS4tAeEm6ZCISPz1eR/Wcy5FX3\n2OzQSlOK6HY+sBZ/OXgJKV1DZ0CGYpqIiUDJXLgAZXvYWed/OujU5y1tFRdd8mLZRGcmk2XB1pyh\n4VJuZkeDRQTbIqimjZJNeKQziF9P9yF/tyuMzEgJD8Ykx/NjWQT95e9dz5GhAp1P6zg5ViqX5sRl\nftHlSgvhwIUsuZn0273vewWvfvJmakaa0wXLTdliPCDjn61cgxFVQcE0kDcMZHUVU7qGxQwVB4DH\n+8LM431hZrHGL6vpdGEqjfbAzSLaD0eH8VBLdb2TadlU0GycTZYwoVpQdRvtrSL+NlnED3tj4LIG\nnlwi46G2ADojAqIy53t/9wD7TmdwdLgIWWCxMRHEc6uieCIRwJEG1YSawcGBXF3F7rnYfWyCtr81\nTF7qy9eyOm1/a9iz57iZuPvZ9cFY3f0s9ji8SrZmKx6fL/mSSaZlE8+xDM+xTFM8QInjoFtOTG1S\nU9EmBaBZNgANBeP2zTWdjbSmoGQaWN3SBssmyuoq3hsZwuqWtqrXqQbhelZDUbdQtBxP4ZdpJ244\nNaEio1r4ja4QOJaBJLAIS4tv+fG5fdwJj8+LhXZWuDWWXl1BrmzYre7aWB4Tme1vDddtUbsVzKXb\nOB9cr8/t9gKa5AHKgsB0hcLM1q6lGCjmMKQUkCwpmNI05HS13GZ2J1AMg4byWZpQiogHguBZlrmS\ny+Dw9Wv43d6VCPA3C1wKqkXpooGxgom8aqGg23iwTcLb4wX8rw+2Q2QYfLs7hLYgj3iER3uYLy/V\nfXzmg+txLkQr71xKJcWwsbZd8oyFuUrezU4UeOGqCd3yHU2zUMXnwbRKBdUqG75K4wc0yQC6fC3m\nHOS4pkK1LWi2jaKpwyCaIRd1u5goFZE3DAR5AV2yUz1+ejIJw7ZnFGsXdQvXJjVopo2CYWNCNfF4\nTwj/86oOXE6pWNsqoTMsIBTgIIscYk1o+Pa5fVTGkM6lVKqnUO3FuZRK+89kaDYZsPngahs20ip2\nYlShE6MKucOYZIGtKhE5MlSgcymV3P5zxbAbFqVNFg3a80mqruioexx7PknRbGGDcymVmqGFuXdb\nN3ZsasfadgnbVoYb9mo1w6azIwolouKsNcNNt+CXpibpzeGrAICVoTC6gyFIHIeuYAgiy6LrNs0L\nTpUUulEsoGQ6xm/NdKzvi8kJpDUVW3tXVB1HQbXow8tZ5DUbX0yUcC2v4/mH2xDmGBy+mEV3WMBD\nbQHIEoveNgmtMo+2Jksy+dwavKbXxWUeTyQCjkyYh+KIO9XNHXPqqqvUdiHUmzVczzs6cCE7QxrL\nVYZJFg3a/fepqviXK/k0m5BqXOZnHTa1c3PHjDGn7giEoknleOTadgm1w+3PpVTqCHLluRxui9vh\nq4XyPOp60wXdz8Hr+WZ2c7gk8wbdyGjY0IDyenOkVCpolQL4ZnsnPkiO4EqxgFbBiVXkDB0hnkeq\npFCLFMB8JbMWgmUTnc+kUDIN8CyLB2Ot4BgGaVXBWEnBb/csr3q9ZlqULpjQLAZF3cJI0cDDXUEE\nAbx3OYfusIAHYiJkiUVLkEc0wCMWbPpH53MbqDR8QLUB64sKeHF9C14/nsb+s1kcGSqU5+PKAosd\nm9pn/Ghd7b+NiSCWRYWGBoXveToB17geGSrQ68fTZUO2bWUYh64UylqOlcbRTQYohg3FsLFlWZi5\nltXJNWTu84NZA2vjUtn4yQKLrSvCODjgzCeuHelZa/yAaikvWWDxwppYuUfY7VF2Pz93YP1gzsBr\nHyXLHnelkvPGRBAnx0rz1oucD5ppUTJngmXQkPEDboEBjAdlxrRsigeC+MWNQXyRn0RCDMIkQp5l\n0SKZKBi6YwjFAGrX5ItBNU26ODWJ4xPO3SYhh7BEjsAiG28PfYm8YeD55Q/OKKS+OKHiSlrD0KSK\noYKB3/laDL+6UUCpZKE7wGNVhecXljh0RGa/cyWLBikVMvkcw0wnhaaP06C6vrfIMmBZIFixXBM4\nFtGAs9xebB3kVxFXDXrn5o4ZsTM3buY+txs3jWLldLzambwurgFx51i4mc+ibpGXgoo7nL2/N4SO\nkMC4S+q4zFeJSLgenzvsuyMklGfmesX/3MSE1/Pu44phlw30Sxtay4YrLvPYvaVzVsGDF9e1lEVi\nK42mO8zcff6lUYVcg7d7Sye21ni8bmvbQsU9XDTDpi+TKmIBblGNEbfEjeE5lkmVFHq+dwX+cvAS\nACAiiDBYDjzrDCjnWQ5pVYFqmlSZiFgoGVWla/mpcu1hQg4hyPHIGxpUy8JAIY/fiHfNGNquaBb9\n8sssVMNGUrUwXjLxNwNT6Ary6Iry6Inc9PyyigWDCO9fypJq2fh44uYP5NqUDoGAdonDv/+V4zW4\nlTtupjjIMih59EnrFkHkGFSGX+ISB8Om8uunNAssAXuPjhLLMhCnNRSjPAOFGCyN8AjyLNrDPIIC\nC4FjYVj2dJE2mnqjuZ9wvRT3R30tq5eXy4upO6snGrt9XQyDOQPPr4pg6/S+XIHa2rKRHZva5xws\nvhBcY+/ur3+pPKfcVT3lZddjdbUVXaPqyo3VluvIArsgZaRKRqY0ShcNrOte/Gdzy9Zx8aBzcFdy\nU/R316/h5FQKbYKEVXBkrguGk7Jv0VVcmpqk6PR8kYgoQWS5WX+wlk2U0UooGDpUy4Rm2bg4lYbI\ncRA5rtySlzN0/L+jwyiYJv7rr61Fi1gtFDmYVunvLkzhzJgC3bQxploQWQaJaW9rSrWQU0t4sC0A\nzSDEZB4RYrCmw9nO5mURSDxTnklyu70zyyZyc0uWTdAtgk0Eywampi/MgqqhoFn4+EqeVMOCZQEF\nw4Ki29Cm+5w100Y04LT1PbREBhhgWYsIjmXAsQx4jrmnax3XxiUcHPB+7pWvx5m9pybJ1UR0l5F7\nnk7gfFrDybHSrP2n7lLVNSau0dx3OoOX3rlB+57pqfrcXE+uctbtnqcTeB03lZxlgUVfVEBSsXAt\nq9Niyz9cde+4zKMjyGEwayBZNMrHMB9pqUoD2d8rl+dzPLc6isHj6bKCtWLY6O8NlT28F9bEGLer\nw41nNqpEXVAtGs/riAZ42ERokWf3+l565wYpho2927qx73QGGxOBukrjtzyQlQiG8EhLOz5IjmDS\n0DBeUhARBHBgIHAcpjQNYUFEhpzK75Jlgmc5TKolqlSddskbOi5l0zBtgkWEkmlAs21ILAuOYTGl\na2iXAjiVnoDAcmAZBt/p6UOrFATHMsxgWqUPB/NIlgg/+XgcmmFDEliIPItvdoqQeBZLIiJ4lkEk\nyEHkGCxpEREP333tbo0cT6WxdAUedItQ0m1olo18yYJqEsamdGimjc+vF6GZNiIShyDP4oOLU9QV\nFTGoWOgOsljdGVjU0uN24no79Ty6lx9rY9wfTVzm0b9URkdIYJyMaAZbV4axq862t69vZQ5cyNLR\n685qYPeWTpwcK2Hf6UzdH7prkPqiArauDJeNy8uPtTHvXM7TG2encD6t4XxaqzK+syU7Kl9Ty87N\nHTg4kMOhKwWkFBPn0xoUI4oX17Xg4EBuVu9vbbs0YxiRo8idKetCVsYfvbzW7etby5/vXFPxKkkX\nDCpN36jdkNB8JOfcpNFrHyUxmDNmHfJ12y7grKZTVi/hP165AAAIsDyWSAF0B0POXOFpgyixLHiW\nhcBy4Cp+39Z0Q0nJNKYNnwkLhJxxU4QhIgg4m89CYFhMTgoI6CFwJEJkGUREDrEAh74WCfGIAEW3\ncCmtYSw37YlKHDpDTqo9Mj0wfWmrhIjENU11+l7GNaCup6noFgqqDc2wkVYMWBYwUTQgiywCHAte\nYJCIigABy1pF8JwjGHG33UR87g4KqkWjWR08xyAscVB0CwLHoLtFavh62XtqkoayOl59qguvH097\nzhlOFg2SBfb2GUDASVIMF3I4mUnhYs6JFXRKQSSkAOSaoeoBrvpOZkyL86nTitMly4Rp2zBsG3nL\nxHjOhqZyyGacO+ZKWUYowGNNPIiWEIelrRKiAQ5BkWMuTCg0kjVwNa0ir1roiYgICSwkngXHAR0R\nEbEAh44of894OXcSbVoAUjdt6BahoFnIKRYmVRNkE9JFE7pFCPCMUz4UcVYAy1pFiDx7zy+xfRbO\neE6nyaIJm4Cw5Hh5AYFdlLjw7mMT5HrYgzkDW1dUe4BHrytIKSb2P9t7Zy46yyZKlYrIGTp+MXIN\nOeNm9ivA8pA5DjFegMA6H4hRIbmfNjVk8gwKORFKkQcn2IiLIr6xNIKlsRC6YoSYFEC9xMrO0xP0\n+4kIzo8rEKZPX5acpEFHWEA4wKIzIvhJg1uI6026S/FJxYRmEJIFA4ZlYzTnePltAR6SxKI7KsIG\n0NcqQeAYPxN+D6IZNiXzBrKqBZFjIPIMNIMQDrBolXkEmzjWc8e7o/T6by1hdh+bILd8CXDktYqm\njW0rboYd7tiFpJkWqZYF1dRxNZ/F4fHrnq9zR29aNgPoAtKpACyDRVeIR29MwuM9IcSCPOIhCQLL\nzCnFXzRs+vhSDpppT/fzMpBFDpHpCXMhkVuU2rTPwqk1jNmSk6ypNIyyyCI2naRKtIogC5BFdsHL\nJZ9by9kRhaYUE50RAUGBRV5zVnC9Daq6u90lcyWEiIjOp7Vy3HdjIoitK8NIKSb2nc6gvzdUNUb0\njlXz1ltaFnSdVMtEpmTiwngJv7yoADyLjQkZS9tEPLCORYDjEOB48Cxb19Orx+nBAkSeQUDkERJZ\nxGQe0QCHsMT5nsUdptFlsGnZZNk3kzrXJzVKFQwUdRtZxURWsxCTOBhECAc4tAZ4yEEWPRGxLGDr\nf+eLQzNs0k0blyZUVH6U0nTSortFxNc6pQWHkoq6RQcHcjh8pTBj1rEXDONcQ26rYF9UwCvvjWHn\n5g6sbZdmbOOuaWdwCxv//ZEJTKgWokEO/UvDePWZeFOSEJpp0cCoCsMmtId5CByDliDvJzjuYRYS\nptAMm2wilAwbBdXC2RGFbMtR/7ZtwkTBgE3ONEDbJkSnVb87wwIYDkiERWRLJnpaxPs6TOLGdUey\nOlTDSXZZRCjoNlgCWoI8eJ5BumggILB4ZKnclM+DiGjf6QxOjpWwY1M7Xj40gv3P9jKHLufp6Dy3\nsf9Mhg5edLpdNiaC6IsKWB4TGSKi779dvdK8KwzgYFqlf/fBCJKGo768sTuM59e3NDUuMDJlQBJY\ntAecZIcssk3dvs+9QT2PrzKRYxFQUC0YNqFQsjBVMjGSccqDTmlOeVCrzOM/n0xRW9gxkPGwABDQ\nKnOISNzNEQkM3ZWJNM20CMTAsglTJRNDkzpYBmiVeai6jc+uF5EvWWgJcRCnY/Eiy+KBdhGSwKJV\nnr1WdyGcS6n0yntjWNsuwVW9ARwRBq9+63oMZg3s2NSOZMlCX1RApdGrrem8Ywbw8PkM/fpGEaLI\n4vT1Iv7Ft7rQIvO3LBu4wp9S5jMLi10KaxWjEA2LUNQtWDag6BYuJ0ukqDZKpg3DIqi67dSxTr8l\nOx0Xc0u9TItg2ASDCLpJiMs8JM67hjFZMqCbhJDAQmAZGDZBmC5ebwvwiAQ4tAR5R+WcUF6mfjmh\nISAwkDgWAs9gTSIIiWea6hS4Hldc5vH8qghOjql4bnW0bhzPNXwvbWjF0eEijg4XsWNTO/adzmDv\ntm4AwMZPUuRV1lLJric7Gbduk4hobbsEIqKXD43MqM287QbwakqlEzeK+OuBKSTCAv75+nYkYn7W\n1efephED6sYua9FNG6pJMLyehGNYndcReI6BNEv4m2OdEM/Nv29fO6Rr+M6nNTy3OoqDF3OIyzxk\ngfUcdFSJYth482IOfVEB59P6tA6g4/29OZCr22I4G25x+J6nE9h/NruoNrwF87PjE7Tr0DD9/NMk\nXW1Ah83Hx+feY/exCdrx7igdGSrQnk9SBHiPuazlyFCBdrw7SkXdou1vDRMRlbdxYlShvacmG7Id\nJ0aVWV9/y6f5fD5SpD/5YITGsjqeXB7GP36kzV+O+vjcpxR1i/afyZDbPtffG8KymOD0Rc8jjre2\nXUJKMXHoamG65c5pveuQOew7nUEjEoJEdFtl+8topkX/9v0b9L+8e50+vJQtS1L7+PjUh4iqBgkR\nER24mJ23wjIR0e5jE1Wvd7c5HyXr3ccmqoYt7T01SdvfGqaX3rnRkHq2u63tbw3TiVGFiIgOXMgS\nEdHeU5Ozno/7GiKiZNEov3/XB2PknsuuD8bowIX5fy6z0fQY4JErWXr9yBi+2RvGoz0htMiLl7ry\n8fkqUCs64OIOO5oP/b0yQh4/Oa8hSrW8uL4FP/pwHNeyOvVFBaQUE1un5ejd+rq5cIUhzqe1smI0\nEdHR6wqeXx3F0eEi1raLs24jLnNgGIYhInpzII+NiSAGcwYUw3ZUckoWDg7kcHAg56m40whNM4CX\nkyV6+2wGjyRC2NwX8ZMaPj7zJFk0KFmysO90BrLAYvtbwxSXeZxPazh4MYf+3tC8txWX+bJqtTsK\nQBZYHLyYqyvS6uLWylUauwMXsvTTU5PY/tYwueo1XtJSRERJxUJniC+rw7w5kMe1rE4MwzB7T03S\nax8l0d8bqitNVfuZvDmQw9HhYlnN5eBADru/nWCInHT5fI3ybDTFAB74NEVvn83g97/e6Xt8Pj4N\nUjQJBy/mkFJMbFsZxpp2CXGZxxtnprC2XcJQVsfeU5NU2cI1H/piIrauDEMxbBwdVuYUPQVuGhVX\nUODgQA6ywGLvtm7IAlvX6LiP7z+Todc+Spbr7VKKiXcu5+mJRACvvDeGHfPQHgQcQ762XcLB6b/7\ne0NlCf5mGD6XRRnAdMGgqZKJvjYJPMcyO5p1VD4+XyGWx0TG9QIPXsxh+/pWpqhbtKZdwgtrYs4S\ncriIl965QdvXxTw9KLeIOKWY2H8mQ7/RG8LRoSJOjpWcZWxMQGV8sVLiHgAmiib98XujZXn8kMgx\nbu3cnqcTODiQw7J5TGZ7cV0LkoqF82mtPIT+fFpD0STIAlseM0BE9ObFnKc0/tFhBS+svvn4/unX\nN9PwuSw4C+yOuXSNX/MOycfnqwURkTtcSTFs7PkkRa8fT6NDvikI3N8bgmLY2Hc6g9qxlEREHUEO\n21aEnSX0+lZGMWycHCvBNapFg7BtZRivPtWFv/ju0irjR0TUIXP4i+8uxdYKpRQA5b/XtEvYdzqD\n2cZmuiyLCdjzdAL9vTJeeW8MBy/mcHC6tu/kWAluKcuQR4b2fFpD/9KZ8vu3wvgBi/AA78b2Hh+f\ne5W17RLWtkuQBRb9vTJkgcVrHyWRLBrlDoa/+O5SADONgfv3jndHyY0hHr5SwHOro/h0rITzKQ3P\nr4pg3+kMtq+f2YXhvp+I6PBVp1C5qFtlj3LXB2O0tl3CnqcTcyo6H75SwOErBVxIadiYCOL131rC\nVCq5EBGdHCuhtifXxWvg063krugF9vH5KuMaoBOjCiVLFl4/nq6aLbKlV8Zv9IZm9YKIiF55bwzb\n18WgGDaGcgYOXsxN19Jp8x7wlFJMuEPW3RGca9ulclZ2rmxy//TwpG0PRBgiohOjCr1xZqpq5ObG\nRBD9vaEFdXU0G9+L8/G5gxAR/ejDcayNS3hxXQu+//b18lD12hKPQ5fztK1i6epyLavT68fT2PN0\nAifHSuXSk5cPjaAvKuC51VGsbXfmA3vN/gUcj++1j5LlchnX03PnHbtcSGtQDBvLYiK2r4vNmlhx\nC5EXO9TpVuJ7gD4+d5hXn+oCwzDMi9PlHS61jftHh4uepSxDWR27t3QCQNVs4R2b2qu8t9mmsckC\ni7jM48V1LQCcZMb5tOYsaa8Wqo6lLyrg5cfamJfncW53s/EDfAPo43NHqVzW1hZCb1sZBioyt4M5\nA0cqhrW7uFlhIqKUYs3Y7p5PUuV6wLmO45U6zy+k9u5WJS6aiW8AfXzuEvqiAvoqlpsvrmvBmnap\nrKDy3OronEmCfo8B5tvXxfD68TQen+6tXQj3gjHz8fHx8YRqltc+Dv8/XOy0BGmbNq4AAAAASUVO\nRK5CYII=\n')
logo = tk.PhotoImage(file="Logo.png")
tk.Label(window, image=logo).place(x=0,y=200) 

window.mainloop() 






