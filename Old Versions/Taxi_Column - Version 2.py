import xlrd
import xlsxwriter
import time
from openpyxl import *

#path = '14decraw.xlsx' #Get filename here

print("Welcome To Rookie441's script.\n")

path = input("Enter File Name: ") + ".xlsx"

file = xlrd.open_workbook(path)

print("Default worksheet is first sheet, named 'Sheet1'.")
answer = input("Is this correct? Y/N: ")

if answer == "Y" or answer == "y":
    sheet = file.sheet_by_index(0) #First sheet in excel file i.e index 0
    wb = load_workbook(path)
    ws = wb["Sheet1"]
else:
    sheet_index = int(input("Which sheet number are u working on? E.g 1,2,3. "))
    sheet_name = input("What is the name of the sheet u are working on? E.g Sheet1, 14dec, 16decbookings, Sheet5. ")
    sheet = file.sheet_by_index(sheet_index-1) #First sheet in excel file i.e index 0
    wb = load_workbook(path)
    ws = wb[sheet_name]

information_list = []
detailed_list = []
driver_list = []
price_list = []
return_list = []
fill_index = 0


for i in range(sheet.nrows):
    #driver
    information_list.append(sheet.cell_value(i,1)) #column here, (i,0) for timing
    s = information_list[i][:4] #checking only first 4 indexes
    for element in (s): 
        if element.isdigit():
            driver_list.append(element)
            if fill_index == 0:
                fill_index = i
            break
    else:
        driver_list.append(" ")
        
    detailed_list.append(information_list[i].split(' '))

for i in range(sheet.nrows):
    try:
        if 'return' in detailed_list[i][0].lower() or "rtn" in detailed_list[i][0].lower() or 'return' in detailed_list[i][1].lower() or "rtn" in detailed_list[i][1].lower():
            return_list.append('yes')
        else:
            return_list.append('no')
    except:
        return_list.append('no')

    #price
    dollar_sign = information_list[i].find('$')
    if "/" in information_list[i] and 'way' in information_list[i] and return_list[i] == 'yes':
        initial_price_index = dollar_sign+1
        price_index = initial_price_index
        while True:
            price_str = information_list[i][price_index]
            if price_str.isdigit():
                price_index+=1
            else:
                break
        if price_index > initial_price_index:
            price_list.append(information_list[i][initial_price_index:price_index]) #string slicing
        else:
            price_list.append(" ")
    elif return_list[i] == 'yes':
        price_list.append("RETURN")
    else:
        price_list.append(" ")

#Edit excel file

for i in range(fill_index,len(information_list)):
    wcell1 = ws.cell(i+1,4)
    wcell1.value = driver_list[i]
    wcell2 = ws.cell(i+1,5)
    wcell2.value = price_list[i]

wb.save(path)

print("File Saved")
print("Success!")
time.sleep(1)
print("...")
time.sleep(1)
print("Program will terminate soon")
time.sleep(1)
print("...")
