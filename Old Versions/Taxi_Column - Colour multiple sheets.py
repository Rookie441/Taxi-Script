import xlrd
import xlsxwriter
import time
import re
from openpyxl import *
from openpyxl.workbook import Workbook
from openpyxl.styles import Font, Fill, PatternFill

def extract_phone_numbers(string):
    r = re.compile(r'(\d{3}[-\.\s]??\d{3}[-\.\s]??\d{4}|\(\d{3}\)\s*\d{3}[-\.\s]??\d{4}|\d{3}[-\.\s]??\d{4})')
    phone_numbers = r.findall(string)
    return [re.sub(r'\D', '', number) for number in phone_numbers]

#Interface
print("Welcome To Rookie441's script.\n")

path = input("Enter File Name: ") + ".xlsx"

file = xlrd.open_workbook(path)

print("Please ensure the name of your sheets are '1','2','3'...\n")
answer = int(input("How many sheets do you have?: "))
    

for j in range(answer):
    sheet = file.sheet_by_index(j) #First sheet in excel file i.e index 0
    wb = load_workbook(path)
    ws = wb[str(j+1)]

    #Lists
    information_list = []
    time_list = []
    split_list = []
    driver_list = []
    price_list = []
    return_list = []
    phone_list = []
    trip_list = []

    #Counters
    total_entries = 0
    first_entry = 0

    #Main
    for i in range(sheet.nrows):
        #time
        if sheet.cell_value(i,0)[0].isdigit():
            time_list.append(sheet.cell_value(i,0))
            if first_entry == 0:
                first_entry = i #first entry same row as first "time" entry

    for i in range(first_entry,sheet.nrows):            
        #driver
        information_list.append(sheet.cell_value(i,1)) #column here, (i,0) for timing
        s = information_list[i-first_entry][:4] #checking only first 4 indexes
        for element in (s): 
            if element.isdigit():
                driver_list.append(element)
                break
        else:
            driver_list.append(" ")
            
        split_list.append(information_list[i-first_entry].split(' '))

    #Debug    
    print(time_list)
    #print(driver_list)
    #print(information_list)
    total_entries = len(information_list)
    #print(total_entries)

    for i in range(len(information_list)): #check elements in first 3 spaces for return/rtn
        try:
            if 'return' in split_list[i][0].lower() or "rtn" in split_list[i][0].lower()\
               or 'return' in split_list[i][1].lower() or "rtn" in split_list[i][1].lower()\
               or 'return' in split_list[i][2].lower() or "rtn" in split_list[i][2].lower():
                return_list.append('yes')
            else:
                return_list.append('no')
        except:
            return_list.append('no')

        #price
        dollar_sign = information_list[i].find('$')
        if "/" in information_list[i] and 'way' in information_list[i] and return_list[i] == 'yes':
            initial_price_index = dollar_sign+1
            price_index = initial_price_index
            while True:
                price_str = information_list[i][price_index]
                if price_str.isdigit():
                    price_index+=1
                else:
                    break
            if price_index > initial_price_index:
                price_list.append(information_list[i][initial_price_index:price_index]) #string slicing
            else:
                price_list.append(" ")
        elif return_list[i] == 'yes':
            price_list.append("RETURN")
        else:
            price_list.append(" ")

        #phone number
        phone_list.append(extract_phone_numbers(information_list[i]))
        
    #trip list by matching phone number
    for i in range(len(phone_list)):  
        if len(phone_list[i]) == 0:
            trip_list.append(0) #no phone number
        else:
            trip_list.append(1) #1 represents yellow, oneway trip
        
            
    for i in range(len(phone_list)):  
        for j in range(len(phone_list)):
            if j != i and phone_list[i] == phone_list[j]:
                if trip_list[i] == 1:
                    trip_list[i] = 2 #2 represents white, start trip of twoway trip
                if trip_list[j] == 1:  
                    trip_list[j] = 3 #3 represents green, end trip of twoway trip
                #elif trip_list[j] == 2: #potential 3rd trip 

    #Edit excel file                    
    column_list = ['A','B','C','D','E','F','G','H','I','J','K','L']
    

    for j in range(len(column_list)):
        c = ws[column_list[j]+str(1)]
        c.fill = PatternFill(patternType='solid',fgColor="5B9BD5") #blue

    for i in range(first_entry,len(information_list)+first_entry):
        wcell1 = ws.cell(i+1,4)
        wcell1.value = driver_list[i-first_entry]
        wcell2 = ws.cell(i+1,5)
        wcell2.value = price_list[i-first_entry]
        if trip_list[i-first_entry] == 0: #gray, no phone number
            for j in range(len(column_list)):
                c = ws[column_list[j]+str(i+1)]
                c.fill = PatternFill(patternType='solid',fgColor="00C0C0C0") #gray
        elif trip_list[i-first_entry] == 1: #yellow
            for j in range(len(column_list)):
                c = ws[column_list[j]+str(i+1)]
                c.fill = PatternFill(patternType='solid',fgColor="00FFFF00") #yellow        
        elif trip_list[i-first_entry] == 3: #green
            for j in range(len(column_list)):
                c = ws[column_list[j]+str(i+1)]
                c.fill = PatternFill(patternType='solid',fgColor="0000FF00") #green         

    wcell1 = ws.cell(total_entries+first_entry+2,4)
    wcell1.value = "Total trips: " + str(total_entries)

    wb.save(path)

print("File Saved")
print("Success!")
time.sleep(1)
print("...")
time.sleep(1)
print("Program will terminate soon")
time.sleep(1)
print("...")

##c = ws['A5']
##c.fill = PatternFill(patternType='solid',fgColor="0000FF00") #green
##c.fill = PatternFill(patternType='solid',fgColor="00FFFF00") #yellow
##c.fill = PatternFill(patternType='solid',fgColor="00C0C0C0") #gray



