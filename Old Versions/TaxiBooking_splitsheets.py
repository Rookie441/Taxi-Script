import xlrd
import xlsxwriter
import time
import re
from openpyxl import *
from openpyxl.workbook import Workbook
from openpyxl.styles import Font, Fill, PatternFill

#Interface
print("Welcome To Rookie441's script.\n")

path = input("Enter File Name: ") + ".xlsx"

file = xlrd.open_workbook(path)

print("Default worksheet is first sheet, named 'Sheet1'.")
answer = input("Is this correct? Y/N: ")

if answer == "Y" or answer == "y":
    sheet = file.sheet_by_index(0) #First sheet in excel file i.e index 0
    wb = load_workbook(path)
    ws = wb["Sheet1"]
else:
    sheet_index = int(input("Which sheet number are u working on? E.g 1,2,3. "))
    sheet_name = input("What is the name of the sheet u are working on? E.g Sheet1, 14dec, 16decbookings, Sheet5. ")
    sheet = file.sheet_by_index(sheet_index-1) #First sheet in excel file i.e index 0
    wb = load_workbook(path)
    ws = wb[sheet_name]

day_list = ['Monday','Tuesday','Wednesday','Thursday','Friday','Saturday','Sunday']
header_list = []

for i in range(sheet.nrows):
    for y in range(7):
        if day_list[y] in sheet.cell_value(i,0).strip(','):
            header_list.append(i)
    try:
        if sheet.cell_value(i,0)[0].isdigit():
            last_entry = i
    except:
        pass
    

working_list = []

for i in range(len(header_list)):
    working_list.append([])
    if i<len(header_list)-1:
        for y in range((header_list[i]+1),(header_list[i+1]-1+1)):
            working_list[i].append(y)
    elif i == len(header_list)-1:
        for y in range((header_list[i]+1),(last_entry+1)):
            working_list[i].append(y)  

workbookOut = xlsxwriter.Workbook("sorted"+path)
cell_format = workbookOut.add_format({'bold':True,'font_color':'white'})

for i in range(len(header_list)):
    worksheetOut = workbookOut.add_worksheet(str(i+1))
    worksheetOut.write("A2",sheet.cell_value(header_list[i],0)) #specific day
    worksheetOut.write("A1","TIME",cell_format)
    worksheetOut.write("B1","DISPATCH",cell_format)
    worksheetOut.write("C1","TRIP FARE",cell_format)
    worksheetOut.write("D1","RETURN DRIVER",cell_format)
    worksheetOut.write("E1","FINAL PAYMENT DUE",cell_format)
    worksheetOut.write("F1","MODE OF PAYMENT",cell_format)
    worksheetOut.write("G1","Albert",cell_format)
    worksheetOut.write("H1","Sirong",cell_format)
    worksheetOut.write("I1","Frankie",cell_format)
    worksheetOut.write("J1","Jay",cell_format)
    worksheetOut.write("K1","James",cell_format)
    worksheetOut.write("L1","CHECK",cell_format)
    counter = 3
    for y in range(len(working_list[i])):
        worksheetOut.write("A"+str(counter),sheet.cell_value(working_list[i][y],0))
        worksheetOut.write("B"+str(counter),sheet.cell_value(working_list[i][y],1))
        counter+=1

print('Success')
print(len(header_list))
workbookOut.close()


