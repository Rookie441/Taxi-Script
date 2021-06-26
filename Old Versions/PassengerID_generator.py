import xlrd
import xlsxwriter
import time
from pyxlsb import convert_date #####
from openpyxl import *
from openpyxl.workbook import Workbook
from openpyxl.styles import Font, Fill, PatternFill

#DEBUG
'''xlrd.__VERSION__''' #pip install xlrd==1.2.0

#Interface
print("Welcome to PassengerID_generator by Rookie441.\n")

path = input("Enter File Name: ") + ".xlsx"

file = xlrd.open_workbook(path) #,formatting_info=True .xls

print("Default worksheet is first sheet, named 'Sheet1'.")
answer = input("Is this correct? Y/N: ")

if answer == "Y" or answer == "y":
    sheet = file.sheet_by_index(0) #First sheet in excel file i.e index 0
    wb = load_workbook(path)
    ws = wb["Sheet1"]
else:
    sheet_index = int(input("Which sheet number are u working on? E.g 1,2,3. "))
    sheet_name = input("What is the name of the sheet u are working on? E.g Sheet1, 14dec, 16decbookings, Sheet5. ")
    sheet = file.sheet_by_index(sheet_index-1) #First sheet in excel file i.e index 0
    wb = load_workbook(path)
    ws = wb[sheet_name]

#Lists
monthToLetter_list = ['A','B','C','D','E','F','G','H','I','J','K','L']
clientID_list = []
work_list = []
month_list = []

def addZero(n):
    if n<10:
        return '0'+str(n)
    else:
        return str(n)
    
def add2Zero(n):
    if n<10:
        return '00'+str(n)
    elif 10<=n<=99:
        return '0'+str(n)
    else:
        return str(n)

def uniqueLetter(name):
    if '#' in name:
        return 'I'
    else:
        return 'R'
    
#Main
for i in range(1,sheet.nrows):  
    ddmmyyyy = format(convert_date(sheet.cell_value(i,9)),'%Y/%m/%d').split('/')
    work_list.append(ddmmyyyy)
    month_list.append(ddmmyyyy[1])
    if i == 1:
        current_month = month_list[0]
        order_counter = 1
    else:
        if month_list[i-1] == current_month:
            order_counter += 1
        else:
            order_counter = 1
            current_month = month_list[i-1]
    order_creation = add2Zero(order_counter)
    clientID = ddmmyyyy[0][2:]+monthToLetter_list[int(ddmmyyyy[1])-1]+order_creation+uniqueLetter(sheet.cell_value(i,1))+add2Zero(i)
    clientID_list.append(clientID)
                                                
#print(clientID_list)
#print(work_list)
print(month_list)

#Debug messages
newID_count = 0
correctID_count = 0
wrongID_count = 0
wrongID_list = []

#Edit excel file
for i in range(len(clientID_list)):
    wcell1 = ws.cell(i+2,1)
    if wcell1.value == None:
        wcell1.value = clientID_list[i]
        newID_count+=1
    elif wcell1.value == clientID_list[i]:
        correctID_count+=1
        pass
    else:
        wcell1.value = clientID_list[i]
        wrongID_count+=1
        wrongID_list.append(i+2)

print()
print('#####')
print(wrongID_count,'of',correctID_count+wrongID_count, 'existing clientID had mistakes and was corrected.')
print('Rows of initially wrong IDs:',wrongID_list)
print()
print('#####')
print(newID_count,'new clientID created.')
print()


wb.save(path)

time.sleep(1)
print("File Saved")
print("Success!")
time.sleep(1)
input("Press ENTER to close the program.")

##c = ws['A5']
##c.fill = PatternFill(patternType='solid',fgColor="0000FF00") #green
##c.fill = PatternFill(patternType='solid',fgColor="00FFFF00") #yellow
##c.fill = PatternFill(patternType='solid',fgColor="00C0C0C0") #gray

#colour wrong ID?
#001 or 01 for ID number?


