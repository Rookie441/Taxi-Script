import xlrd
import xlsxwriter
import time
import re
from pyxlsb import convert_date
from openpyxl import *
from openpyxl.workbook import Workbook
from openpyxl.styles import Font, Fill, PatternFill

'''
Changelog v1.1
- Added total trips calculation for each day into every sheet when running option 2.
- Added new headers for summary sheet, including entire sheet total trips, prices and days selected.
- Optimized the code such that user does not need to input sheet name when loading file.
- Fixed a bug causing the prices to lose its number format.
- Added an error message disallowing option 2 to run if there are sheets of the incorrect format.
- New sheets created are now renamed to the exact date instead of '1','2','3'.
'''

#Interface
print("Welcome To Rookie441's script.\n")
print("1) Split information into 1 or multiple sheet(s) and colour them accordingly.\n2) Tabulate total price for 1 or multiple day(s) and collate into a new sheet, appended at the back.\n3) PassengerID Generator")
option = '0'
while not(option=='1' or option=='2' or option=='3'):
    option = input("Enter your Choice: ")

if option == '1':
    while True:
        try:
            path = input("Enter File Name: ") + ".xlsx"
            file = xlrd.open_workbook(path)
        except:
            continue        
        break

    while True:
        try:
            print("You are working on sheet number 1")
            answer = input("Is this correct? Y/N: ")

            if answer == "Y" or answer == "y":
                sheet = file.sheet_by_index(0) #First sheet in excel file i.e index 0
                wb = load_workbook(path)
                ws = wb.worksheets[0]
            elif answer == "N" or answer == "n":
                sheet_index = int(input("Which sheet number are u working on? E.g 1,2,3. "))
                sheet = file.sheet_by_index(sheet_index-1) #First sheet in excel file i.e index 0
                wb = load_workbook(path)
                ws = wb.worksheets[sheet_index-1]
            else:
                continue
        except:
            continue        
        break
    

    day_list = ['Monday','Tuesday','Wednesday','Thursday','Friday','Saturday','Sunday']
    header_list = []

    for i in range(sheet.nrows):
        for y in range(7):
            if day_list[y] in sheet.cell_value(i,0).strip(','):
                header_list.append(i)
        try:
            if sheet.cell_value(i,0)[0].isdigit():
                last_entry = i
        except:
            pass
        

    working_list = []

    for i in range(len(header_list)):
        working_list.append([])
        if i<len(header_list)-1:
            for y in range((header_list[i]+1),(header_list[i+1]-1+1)):
                working_list[i].append(y)
        elif i == len(header_list)-1:
            for y in range((header_list[i]+1),(last_entry+1)):
                working_list[i].append(y)  

    workbookOut = xlsxwriter.Workbook("sorted"+path)
    cell_format = workbookOut.add_format({'bold':True,'font_color':'white'})

    startdate_list = []
    for i in range(len(header_list)):
        worksheetOut = workbookOut.add_worksheet(sheet.cell_value(header_list[i],0)) ####c = ws[str(sheet.cell_value(1,0))] ########RENAME FILE HERE########
        worksheetOut.write("A2",sheet.cell_value(header_list[i],0)) #specific day
        startdate_list.append(sheet.cell_value(header_list[i],0))
        worksheetOut.write("A1","TIME",cell_format)
        worksheetOut.write("B1","DISPATCH",cell_format)
        worksheetOut.write("C1","TRIP FARE",cell_format)
        worksheetOut.write("D1","RETURN DRIVER",cell_format)
        worksheetOut.write("E1","FINAL PAYMENT DUE",cell_format)
        worksheetOut.write("F1","MODE OF PAYMENT",cell_format)
        worksheetOut.write("G1","Albert",cell_format)
        worksheetOut.write("H1","Sirong",cell_format)
        worksheetOut.write("I1","Frankie",cell_format)
        worksheetOut.write("J1","Jay",cell_format)
        worksheetOut.write("K1","James",cell_format)
        worksheetOut.write("L1","CHECK",cell_format)
        counter = 3
        for y in range(len(working_list[i])):
            worksheetOut.write("A"+str(counter),sheet.cell_value(working_list[i][y],0))
            worksheetOut.write("B"+str(counter),sheet.cell_value(working_list[i][y],1))
            counter+=1

    print('Success splitting into',len(header_list),'sheets')
    #print(startdate_list)
    workbookOut.close()


    #Colour Multiple Sheets
    print('')
    def extract_phone_numbers(string):
        r = re.compile(r'(\d{3}[-\.\s]??\d{3}[-\.\s]??\d{4}|\(\d{3}\)\s*\d{3}[-\.\s]??\d{4}|\d{3}[-\.\s]??\d{4})')
        phone_numbers = r.findall(string)
        return [re.sub(r'\D', '', number) for number in phone_numbers]

    print('Colouring multiple sheets')

    path = "sorted"+path
    file = xlrd.open_workbook(path)

    for j in range(len(header_list)):
        sheet = file.sheet_by_index(j) #First sheet in excel file i.e index 0
        wb = load_workbook(path)
        ws = wb.worksheets[j]

        #Lists
        information_list = []
        time_list = []
        split_list = []
        driver_list = []
        price_list = []
        return_list = []
        phone_list = []
        trip_list = []
        #calculate_price_list = []

        #Counters
        total_entries = 0
        first_entry = 0

        #Main
        for i in range(sheet.nrows):
            #time
            if sheet.cell_value(i,0)[0].isdigit():
                time_list.append(sheet.cell_value(i,0))
                if first_entry == 0:
                    first_entry = i #first entry same row as first "time" entry

        for i in range(first_entry,sheet.nrows):            
            #driver
            information_list.append(sheet.cell_value(i,1)) #column here, (i,0) for timing
            s = information_list[i-first_entry][:4] #checking only first 4 indexes
            for element in (s): 
                if element.isdigit():
                    driver_list.append(element)
                    break
            else:
                driver_list.append(" ")
                
            split_list.append(information_list[i-first_entry].split(' '))


        #Debug
        print(str(j+1),"...")
        #print(driver_list)
        #print(information_list)
        total_entries = len(information_list)
        #print(total_entries)

        for i in range(len(information_list)): #check elements in first 3 spaces for return/rtn
            try:
                if 'return' in split_list[i][0].lower() or "rtn" in split_list[i][0].lower()\
                   or 'return' in split_list[i][1].lower() or "rtn" in split_list[i][1].lower()\
                   or 'return' in split_list[i][2].lower() or "rtn" in split_list[i][2].lower():
                    return_list.append('yes')
                else:
                    return_list.append('no')
            except:
                return_list.append('no')

            #price
            dollar_sign = information_list[i].find('$')
            if "/" in information_list[i] and 'way' in information_list[i] and return_list[i] == 'yes':
                initial_price_index = dollar_sign+1
                price_index = initial_price_index
                while True:
                    price_str = information_list[i][price_index]
                    if price_str.isdigit():
                        price_index+=1
                    else:
                        break
                if price_index > initial_price_index:
                    price_list.append(int(information_list[i][initial_price_index:price_index])) #string slicing
                else:
                    price_list.append(" ")
            elif return_list[i] == 'yes':
                price_list.append("RETURN")
            else:
                price_list.append(" ")

            #phone number
            phone_list.append(extract_phone_numbers(information_list[i]))

            
        #trip list by matching phone number
        for i in range(len(phone_list)):  
            if len(phone_list[i]) == 0:
                trip_list.append(0) #no phone number
            else:
                trip_list.append(1) #1 represents yellow, oneway trip
            
                
        for i in range(len(phone_list)):  
            for j in range(len(phone_list)):
                if j != i and phone_list[i] == phone_list[j]:
                    if trip_list[i] == 1:
                        trip_list[i] = 2 #2 represents white, start trip of twoway trip
                    if trip_list[j] == 1:  
                        trip_list[j] = 3 #3 represents green, end trip of twoway trip
                    #elif trip_list[j] == 2: #potential 3rd trip

        #Edit excel file                    
        column_list = ['A','B','C','D','E','F','G','H','I','J','K','L']
        

        for j in range(len(column_list)):
            c = ws[column_list[j]+str(1)] 
            c.fill = PatternFill(patternType='solid',fgColor="5B9BD5") #blue

        for i in range(first_entry,len(information_list)+first_entry):
            wcell1 = ws.cell(i+1,4)
            wcell1.value = driver_list[i-first_entry]
            wcell2 = ws.cell(i+1,5)
            wcell2.value = price_list[i-first_entry]
            if trip_list[i-first_entry] == 0: #gray, no phone number
                for j in range(len(column_list)):
                    c = ws[column_list[j]+str(i+1)]
                    c.fill = PatternFill(patternType='solid',fgColor="00C0C0C0") #gray
            elif trip_list[i-first_entry] == 1: #yellow, one way
                for j in range(len(column_list)):
                    c = ws[column_list[j]+str(i+1)]
                    c.fill = PatternFill(patternType='solid',fgColor="00FFFF00") #yellow        
            elif trip_list[i-first_entry] == 3: #green, 2nd of 2 way
                for j in range(len(column_list)):
                    c = ws[column_list[j]+str(i+1)]
                    c.fill = PatternFill(patternType='solid',fgColor="0000FF00") #green         

        
        wcell1 = ws.cell(total_entries+first_entry+2,4)
        wcell1.value = "Total trips: " + str(total_entries)

        wb.save(path)
        
        #print(calculate_price_list)

        #column E final payment

elif option == '2':
    while True:
        try:
            path = input("Enter File Name: ") + ".xlsx"
            wb = load_workbook(path)
            numberOfSheets = len(wb.sheetnames)
        except:
            continue        
        break

    total_price = 0
    total_price_list = []
    date_list = []
    total_trips = 0
    total_trips_list = []
    
    for i in range(numberOfSheets):
        ws = wb.worksheets[i]
        if ws.cell(1,1).value == "Date":
            print("ERROR! Summary file detected. Please delete the Summary sheet before executing the program.")
            time.sleep(1)
            input("Press ENTER to close the program.")
            exit()
        elif ws.cell(1,1).value != "TIME":
            print("ERROR! File contained a sheet in the incorrect format")
            time.sleep(1)
            input("Press ENTER to close the program.")
            exit()
         
        date_list.append(ws.cell(2,1).value)
        for k in range(ws.max_row):
            try:
                if ws.cell(k+1,1).value[0].isdigit():
                    total_trips+=1
            except:
                pass
        #print(total_trips)
        total_trips_list.append(total_trips)
        
        for j in range(total_trips+1): 
            try:
                wcellprice = ws.cell(j+2,5)
                total_price+=int((wcellprice.value))
            except:
                pass
        total_price_list.append(total_price)
        wcellprice_day = ws.cell(total_trips+4,5)
        wcellprice_day.value = "Total price: " + str(total_price)
        total_price = 0
        total_trips = 0 #order matters
            
        
    #print(total_trips)
    #print(date_list)
    #Editing Excel 
    wb.create_sheet("Summary")
    ws = wb["Summary"]
    numberOfDates = len(date_list)
    ft = Font(bold=True,color='FFFFFF')
    column_list = ['A','B','C','D']
    column_list_header = ['Date','Total Trips','Total Price','Total days selected']
    for j in range(3):
        c = ws[column_list[j]+str(1)]
        c.fill = PatternFill(patternType='solid',fgColor="5B9BD5") #blue
        c.font = ft
        wcellheader = ws.cell(1,j+1)
        wcellheader.value = column_list_header[j]

    for i in range(numberOfDates):
        wcelldate = ws.cell(i+2,1)
        wcelldate.value = date_list[i]

        wcelltotaltrips = ws.cell(i+2,2)
        wcelltotaltrips.value = total_trips_list[i]
        
        wcelltotalprice = ws.cell(i+2,3)
        wcelltotalprice.value = total_price_list[i]

    for j in range(4):
        c = ws[column_list[j]+str(numberOfDates+4)]
        c.fill = PatternFill(patternType='solid',fgColor="5B9BD5") #blue
        c.font = ft
        wcellheader = ws.cell(numberOfDates+4,j+1)
        wcellheader.value = column_list_header[j]

    wcell_summary = ws.cell(numberOfDates+5,1)
    wcell_summary.value = "Total Summary"
    wcelltotaltrips_summary = ws.cell(numberOfDates+5,2)
    wcelltotaltrips_summary.value = sum(total_trips_list)
    wcelltotalprice_summary = ws.cell(numberOfDates+5,3)
    wcelltotalprice_summary.value = sum(total_price_list)
    wcelltotaldays_summary = ws.cell(numberOfDates+5,4)
    wcelltotaldays_summary.value = numberOfDates

    wb.save(path)
    
elif option == '3':
    while True:
        try:
            path = input("Enter File Name: ") + ".xlsx"
            file = xlrd.open_workbook(path)
        except:
            continue        
        break

    while True:
        try:
            print("You are working on sheet number 1")
            answer = input("Is this correct? Y/N: ")

            if answer == "Y" or answer == "y":
                sheet = file.sheet_by_index(0) #First sheet in excel file i.e index 0
                wb = load_workbook(path)
                ws = wb.worksheets[0]
            elif answer == "N" or answer == "n":
                sheet_index = int(input("Which sheet number are u working on? E.g 1,2,3. "))
                sheet = file.sheet_by_index(sheet_index-1) #First sheet in excel file i.e index 0
                wb = load_workbook(path)
                ws = wb.worksheets[sheet_index-1]
            else:
                continue
        except:
            continue        
        break

    #Lists
    monthToLetter_list = ['A','B','C','D','E','F','G','H','I','J','K','L']
    clientID_list = []
    work_list = []
    month_list = []
        
    def add2Zero(n):
        if n<10:
            return '00'+str(n)
        elif 10<=n<=99:
            return '0'+str(n)
        else:
            return str(n)

    def uniqueLetter(name):
        if '#' in name:
            return 'I'
        else:
            return 'R'
        
    #Main
    for i in range(1,sheet.nrows):  
        ddmmyyyy = format(convert_date(sheet.cell_value(i,9)),'%Y/%m/%d').split('/')
        work_list.append(ddmmyyyy)
        month_list.append(ddmmyyyy[1])
        if i == 1:
            current_month = month_list[0]
            order_counter = 1
        else:
            if month_list[i-1] == current_month:
                order_counter += 1
            else:
                order_counter = 1
                current_month = month_list[i-1]
        order_creation = add2Zero(order_counter)
        clientID = ddmmyyyy[0][2:]+monthToLetter_list[int(ddmmyyyy[1])-1]+order_creation+uniqueLetter(sheet.cell_value(i,1))+add2Zero(i)
        clientID_list.append(clientID)
                                                    
    #print(clientID_list)
    #print(work_list)
    #print(month_list)

    #Debug messages
    newID_count = 0
    correctID_count = 0
    wrongID_count = 0
    wrongID_list = []

    #Edit excel file
    for i in range(len(clientID_list)):
        wcell1 = ws.cell(i+2,1)
        if wcell1.value == None:
            wcell1.value = clientID_list[i]
            newID_count+=1
        elif wcell1.value == clientID_list[i]:
            correctID_count+=1
            pass
        else:
            wcell1.value = clientID_list[i]
            wrongID_count+=1
            wrongID_list.append(i+2)

    print()
    print('#####')
    print(wrongID_count,'of',correctID_count+wrongID_count, 'existing clientID had mistakes and was corrected.')
    print('Rows of initially wrong IDs:',wrongID_list)
    print()
    print('#####')
    print(newID_count,'new clientID created.')
    print()


    wb.save(path)
    

time.sleep(1)
print("File Saved")
print("Success!")
time.sleep(1)
input("Press ENTER to close the program.")



