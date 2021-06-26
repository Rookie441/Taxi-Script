import xlrd
import xlsxwriter
import time
from pyxlsb import convert_date
from openpyxl import *
from openpyxl.workbook import Workbook
from openpyxl.styles import Font, Fill, PatternFill

#Interface
print("Welcome To Rookie441's script.\n")

while True:
    try:
        path = input("Enter File Name: ") + ".xlsx"
        file = xlrd.open_workbook(path)
    except:
        continue        
    break

while True:
    try:
        print("You are working on sheet number 1")
        answer = input("Is this correct? Y/N: ")

        if answer == "Y" or answer == "y":
            sheet = file.sheet_by_index(0) #First sheet in excel file i.e index 0
            wb = load_workbook(path)
            ws = wb.worksheets[0]
        elif answer == "N" or answer == "n":
            sheet_index = int(input("Which sheet number are u working on? E.g 1,2,3. "))
            sheet = file.sheet_by_index(sheet_index-1) #First sheet in excel file i.e index 0
            wb = load_workbook(path)
            ws = wb.worksheets[sheet_index-1]
        else:
            continue
    except:
        continue        
    break

day_list = ['Monday','Tuesday','Wednesday','Thursday','Friday','Saturday','Sunday'] #Must be Caps first word
header_list = []

for i in range(sheet.nrows):
    for y in range(7):
        if day_list[y] in sheet.cell_value(i,0).strip(','):
            header_list.append(i)
        if len(sheet.cell_value(i,0)) >= 8:
            last_entry = i
            
working_list = []

for i in range(len(header_list)):
    working_list.append([])
    if i<len(header_list)-1:
        for y in range((header_list[i]),(header_list[i+1]-1+1)):
            working_list[i].append(y)
    elif i == len(header_list)-1:
        for y in range((header_list[i]),(last_entry+1)):
            working_list[i].append(y)  

#Clear
for i in range(sheet.nrows):
    ws.cell(i+1,1).value = ''

header_counter = 1
time_counter = 2
trip_counter = 2

for i in range(len(header_list)):
    
    entry_row_count = len(working_list[i])-1
    ws.cell(header_counter,1).value = sheet.cell_value(working_list[i][0],0)
    
    for y in range(1,entry_row_count,2):
        ws.cell(time_counter,1).value = sheet.cell_value(working_list[i][y],0)
        time_counter+=1
        
        
    for y in range(2,entry_row_count+1,2):
        ws.cell(trip_counter,2).value = sheet.cell_value(working_list[i][y],0)
        trip_counter+=1

    header_counter = time_counter
    time_counter+=1
    trip_counter+=1




wb.save(path)
    
time.sleep(1)
print("File Saved")
print("Success!")
time.sleep(1)
input("Press ENTER to close the program.")



